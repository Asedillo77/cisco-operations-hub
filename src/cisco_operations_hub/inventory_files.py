from __future__ import annotations

import csv
import shutil
from collections.abc import Iterator
from contextlib import contextmanager
from pathlib import Path
from uuid import uuid4

from openpyxl import load_workbook

PROJECT_ROOT = Path(__file__).resolve().parents[2]
RUNTIME_ROOT = PROJECT_ROOT / ".runtime"


@contextmanager
def prepare_tabular_inventory(path: Path) -> Iterator[Path]:
    """Yield CSV directly or create a short-lived CSV representation of an XLSX file."""
    source = path.expanduser().resolve()
    if not source.is_file():
        raise FileNotFoundError(f"Inventory file was not found: {source}")
    if source.suffix.lower() == ".csv":
        yield source
        return
    if source.suffix.lower() != ".xlsx":
        raise ValueError("Inventory file must be CSV or XLSX.")

    work_dir = RUNTIME_ROOT / f"inventory_{uuid4().hex}"
    work_dir.mkdir(parents=True, exist_ok=False)
    prepared = work_dir / "inventory.csv"
    try:
        _xlsx_to_csv(source, prepared)
        yield prepared
    finally:
        shutil.rmtree(work_dir, ignore_errors=True)


def _xlsx_to_csv(source: Path, destination: Path) -> None:
    workbook = load_workbook(source, read_only=True, data_only=True)
    try:
        sheet = workbook.active
        rows = sheet.iter_rows(values_only=True)
        raw_headers = next(rows, None)
        if not raw_headers:
            raise ValueError("Inventory workbook has no header row.")
        headers = [_cell_text(value) for value in raw_headers]
        if not any(headers):
            raise ValueError("Inventory workbook has no named columns.")
        populated = [header.lower() for header in headers if header]
        if len(populated) != len(set(populated)):
            raise ValueError("Inventory workbook contains duplicate column names.")
        with destination.open("w", encoding="utf-8-sig", newline="") as handle:
            writer = csv.writer(handle)
            writer.writerow(headers)
            writer.writerows([_cell_text(value) for value in row] for row in rows)
    finally:
        workbook.close()


def _cell_text(value: object) -> str:
    return "" if value is None else str(value).strip()
