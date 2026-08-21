from __future__ import annotations

import csv
from collections.abc import Iterable
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

from .models import DeviceTarget

SUPPORTED_DEVICE_TYPES = {"switch", "edge_router"}


def load_inventory(path: Path, max_devices: int = 50) -> list[DeviceTarget]:
    if not path.is_file():
        raise FileNotFoundError(f"Inventory file was not found: {path}")
    if max_devices < 1:
        raise ValueError("Maximum devices must be at least 1.")

    suffix = path.suffix.lower()
    if suffix == ".csv":
        rows = _read_csv(path)
    elif suffix == ".xlsx":
        rows = _read_xlsx(path)
    else:
        raise ValueError("Inventory file must be CSV or XLSX.")

    devices = _normalize_rows(rows)
    if not devices:
        raise ValueError("Inventory contains no enabled device rows.")
    if len(devices) > max_devices:
        raise ValueError(
            f"Inventory contains {len(devices)} enabled devices; the limit is {max_devices}."
        )
    return devices


def _read_csv(path: Path) -> list[tuple[int, dict[str, Any]]]:
    with path.open("r", encoding="utf-8-sig", newline="") as handle:
        reader = csv.DictReader(handle)
        if not reader.fieldnames:
            raise ValueError("Inventory CSV has no header row.")
        return [(number, dict(row)) for number, row in enumerate(reader, start=2)]


def _read_xlsx(path: Path) -> list[tuple[int, dict[str, Any]]]:
    workbook = load_workbook(path, read_only=True, data_only=True)
    try:
        sheet = workbook.active
        values = sheet.iter_rows(values_only=True)
        headers = next(values, None)
        if not headers:
            raise ValueError("Inventory workbook has no header row.")
        names = [str(value).strip() if value is not None else "" for value in headers]
        return [
            (number, dict(zip(names, row, strict=False)))
            for number, row in enumerate(values, start=2)
        ]
    finally:
        workbook.close()


def _normalize_rows(rows: Iterable[tuple[int, dict[str, Any]]]) -> list[DeviceTarget]:
    devices: list[DeviceTarget] = []
    errors: list[str] = []
    seen: set[tuple[str, str]] = set()
    for row_number, row in rows:
        normalized = {_key(key): _value(value) for key, value in row.items() if key is not None}
        if not any(normalized.values()):
            continue
        enabled_value = _first(normalized, "enabled", "include", "active") or "true"
        try:
            enabled = _parse_bool(enabled_value)
        except ValueError as exc:
            errors.append(f"Row {row_number}: {exc}")
            continue
        if not enabled:
            continue
        hostname = _first(normalized, "hostname", "device_name", "name")
        ip_address = _first(normalized, "ip_address", "ip", "management_ip", "host")
        device_type = (_first(normalized, "device_type", "type", "platform") or "switch").lower()
        if not hostname and not ip_address:
            errors.append(f"Row {row_number}: hostname or IP address is required.")
            continue
        if device_type not in SUPPORTED_DEVICE_TYPES:
            errors.append(
                f"Row {row_number}: device_type must be switch or edge_router, not {device_type!r}."
            )
            continue
        hostname = hostname or ip_address
        ip_address = ip_address or hostname
        identity = (hostname.lower(), ip_address.lower())
        if identity in seen:
            errors.append(f"Row {row_number}: duplicate device {hostname} ({ip_address}).")
            continue
        seen.add(identity)
        devices.append(DeviceTarget(hostname, ip_address, device_type, row_number, enabled))
    if errors:
        raise ValueError("Inventory validation failed:\n" + "\n".join(errors))
    return devices


def _key(value: Any) -> str:
    return str(value).strip().lower().replace(" ", "_").replace("-", "_")


def _value(value: Any) -> str:
    return "" if value is None else str(value).strip()


def _first(data: dict[str, str], *names: str) -> str:
    return next((data[name] for name in names if data.get(name)), "")


def _parse_bool(value: str) -> bool:
    normalized = value.strip().lower()
    if normalized in {"true", "yes", "y", "1", "enabled"}:
        return True
    if normalized in {"false", "no", "n", "0", "disabled"}:
        return False
    raise ValueError(f"enabled value {value!r} is not true or false.")
